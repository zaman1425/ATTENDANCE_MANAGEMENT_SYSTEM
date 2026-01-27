from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify, send_file
import io
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
from psycopg2.extras import DictCursor, RealDictCursor
from datetime import date, datetime
import os
from dotenv import load_dotenv
from zipfile import BadZipFile
from openpyxl.styles import Font


app = Flask(__name__)
app.secret_key = "super-secret-key"

load_dotenv()

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=os.getenv("DB_PORT", 5432),
        cursor_factory=DictCursor
    )


def ensure_admin_requests_table():
    """Create the admin_requests table if it doesn't exist yet."""
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_requests (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT now()
            )
            """
        )
        conn.commit()
        cur.close()
    except Exception:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
    finally:
        if conn:
            conn.close()
    
@app.route("/admin_signup_success")
def admin_signup_success():
    return render_template("admin_signup_success.html")


@app.route("/")
def login():
    return render_template("login.html")

@app.route("/admin_signup", methods=["GET", "POST"])
def admin_signup():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        confirm = request.form["confirm_password"]

        if password != confirm:
            return render_template(
                "admin_signup.html",
                error="Passwords do not match"
            )

        hashed = generate_password_hash(password)

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT id FROM admins WHERE email=%s", (email,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return render_template(
                "admin_signup.html",
                error="Admin already exists. Please login."
            )

        # If there's already a pending request, inform user
        cur.execute("SELECT id FROM admin_requests WHERE email=%s", (email,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return render_template(
                "admin_signup.html",
                error="A signup request for this email is already pending approval."
            )

        # Insert into pending admin requests for approval by an existing admin
        cur.execute(
            "INSERT INTO admin_requests (email, password_hash, created_at) VALUES (%s, %s, %s)",
            (email, hashed, datetime.now())
        )
        conn.commit()
        cur.close()
        conn.close()
        flash("Your admin signup request has been submitted for approval.", "info")
        return redirect(url_for("admin_signup_success"))

    return render_template("admin_signup.html")


@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        next_target = request.form.get("next", "")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT password FROM admins WHERE email=%s", (email,))
        admin = cur.fetchone()
        cur.close()
        conn.close()

        if not admin or not check_password_hash(admin[0], password):
            return render_template(
                "admin_login.html",
                error="Invalid credentials",
                next=next_target
            )

        session["admin_logged_in"] = True
        session["admin_email"] = email
        if next_target == 'requests':
            return redirect(url_for('admin_requests'))

        return redirect(url_for("admin_dashboard"))
    next_target = request.args.get('next', '')
    return render_template("admin_login.html", next=next_target)


@app.route('/admin_requests')
def admin_requests():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, email, created_at FROM admin_requests ORDER BY created_at DESC")
    reqs = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('admin_requests.html', requests=reqs)


@app.route('/admin_requests/approve/<int:req_id>', methods=['POST'])
def admin_request_approve(req_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT id, email, password_hash FROM admin_requests WHERE id = %s', (req_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash('Request not found', 'error')
        return redirect(url_for('admin_requests'))
    try:
        cur.execute('INSERT INTO admins (email, password) VALUES (%s, %s)', (row['email'], row['password_hash']))
        cur.execute('DELETE FROM admin_requests WHERE id = %s', (req_id,))
        conn.commit()
        flash(f"Approved admin: {row['email']}", 'success')
    except Exception as e:
        conn.rollback()
        flash('Error approving request: ' + str(e), 'error')

    cur.close()
    conn.close()
    return redirect(url_for('admin_requests'))


@app.route('/admin_requests/reject/<int:req_id>', methods=['POST'])
def admin_request_reject(req_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM admin_requests WHERE id = %s', (req_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Request rejected and removed.', 'info')
    return redirect(url_for('admin_requests'))



@app.route("/admin_dashboard", methods=["GET", "POST"])
def admin_dashboard():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    interns = []
    count = 0
    reg_no = request.values.get("reg_no", "").strip()
    intern_name = request.values.get("intern_name", "").strip()
    reg_error = None
    name_error = None
    general_error = None
    if request.method == "POST":
        if not reg_no and not intern_name:
            reg_error = "Please provide Reg No or Intern Name."
            name_error = "Please provide Reg No or Intern Name."

        else:
            conn = get_db_connection()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            if reg_no and intern_name:
                cur.execute(
                    "SELECT reg_no, intern_name, email FROM interns WHERE CAST(reg_no AS TEXT) = %s AND intern_name ILIKE %s ORDER BY reg_no",
                    (reg_no, f"%{intern_name}%")
                )
            elif reg_no:
                cur.execute(
                    "SELECT reg_no, intern_name, email FROM interns WHERE CAST(reg_no AS TEXT) = %s ORDER BY reg_no",
                    (reg_no,)
                )
            else:
                cur.execute(
                    "SELECT reg_no, intern_name, email FROM interns WHERE intern_name ILIKE %s ORDER BY reg_no",
                    (f"%{intern_name}%",)
                )

            interns = cur.fetchall()
            count = len(interns)

            if count == 0:
                general_error = "No intern found matching the provided search criteria."

            cur.close()
            conn.close()
    return render_template(
        "admin_dashboard.html",
        ints=interns,
        cnt=count,
        reg_no=reg_no,
        intern_name=intern_name,
        reg_error=reg_error,
        name_error=name_error,
        general_error=general_error
    )

@app.route("/admin_attendance/<reg_no>", methods=["GET", "POST"])
def admin_attendance(reg_no):
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=DictCursor)

    cur.execute(
        "SELECT reg_no, intern_name, email FROM interns WHERE reg_no = %s",
        (reg_no,)
    )
    intern = cur.fetchone()

    if not intern:
        cur.close()
        conn.close()
        return "Intern not found", 404

    today = date.today()
    cur.execute(
        """
        SELECT status
        FROM attendance
        WHERE intern_email = %s AND date = %s
        """,
        (intern["email"], today)
    )
    today_status = cur.fetchone()
    if request.method == "POST":
        status = request.form.get("status")  

        if status not in ("P", "A"):
            cur.close()
            conn.close()
            return redirect(url_for("admin_attendance", reg_no=reg_no))

        if today_status:
            cur.execute(
                """
                UPDATE attendance
                SET status = %s, marked_at = %s
                WHERE intern_email = %s AND date = %s
                """,
                (status, datetime.now(), intern["email"], today)
            )
            flash("Attendance updated for today.", "success")
        else:
            cur.execute(
                """
                INSERT INTO attendance (intern_email, date, status, marked_at)
                VALUES (%s, %s, %s, %s)
                """,
                (intern["email"], today, status, datetime.now())
            )
            flash("Attendance recorded for today.", "success")

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("admin_attendance", reg_no=reg_no))

    cur.execute(
        """
        SELECT date, status, marked_at
        FROM attendance
        WHERE intern_email = %s
        ORDER BY date DESC
        """,
        (intern["email"],)
    )
    history = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "admin_attendance.html",
        intern=intern,
        today_status=today_status["status"] if today_status else None,
        attendance=history
    
    )

@app.route("/intern_attendance")
def intern_attendance():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))

    return render_template("intern_attendance.html")


@app.route("/success/<reg_no>")
def success(reg_no):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM interns WHERE reg_no = %s",
        (reg_no,)
    )
    intern = cur.fetchone()

    cur.close()
    conn.close()

    if not intern:
        return "Intern not found", 404

    return render_template("success.html", intern=intern)

@app.route("/already", methods=["GET", "POST"])
def already():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT password_hash FROM interns WHERE email = %s",
            (email,)
        )
        intern = cur.fetchone()
       
        cur.close()
        conn.close()

        if intern and check_password_hash(intern["password_hash"], password):
            session["intern_logged_in"] = True
            session["intern_email"] = email
            return redirect(url_for("intern_dashboard"))

        return render_template("already.html", error="Invalid intern credentials")

    return render_template("already.html")



@app.route("/home", methods=["GET", "POST"])
def home():
    if request.method == "POST":

        reg_no = request.form.get("reg_no")
        intern_name = request.form.get("name")
        age = request.form.get("age")
        contact = request.form.get("contact")
        college = request.form.get("college")
        course = request.form.get("course")
        duration = int(request.form.get("duration"))
        reference_by = request.form.get("reference")
        project = request.form.get("project")
        email = request.form.get("email")
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")

        if password != confirm_password:
            flash("Passwords do not match!", "error")
            return render_template("home.html", reg_no=reg_no, intern_name=intern_name, age=age,
                                   contact=contact, college=college, course=course,
                                   duration=duration, reference_by=reference_by,
                                   project=project, email=email)
        password_hash = generate_password_hash(
            password,
            method="pbkdf2:sha256",
            salt_length=16
        )
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute("""
            INSERT INTO interns
            (reg_no, intern_name, age, contact, college, course,
             duration, reference_by, project, email, password_hash)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING reg_no
        """, (
            reg_no,
            intern_name,
            age,
            contact,
            college,
            course,
            int(duration),
            reference_by,
            project,
            email,
            password_hash
        ))

        reg_no = cur.fetchone()["reg_no"]

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("success", reg_no=reg_no))

    return render_template("home.html")

from flask import flash

@app.route("/intern_dashboard", methods=["GET", "POST"])
def intern_dashboard():
    if not session.get("intern_logged_in"):
        return redirect(url_for("already"))

    email = session["intern_email"]

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM interns WHERE email = %s",
        (email,)
    )
    intern = cur.fetchone()

    if not intern:
        cur.close()
        conn.close()
        return "Intern not found", 404
    cur.execute(
        "SELECT date, status FROM attendance WHERE intern_email = %s ORDER BY date DESC",
        (email,)
    )
    attendance = cur.fetchall()
    if request.method == "POST":
        new_data = {
            "intern_name": request.form.get("intern_name", "").strip(),
            "age": request.form.get("age", "").strip(),
            "contact": request.form.get("contact", "").strip(),
            "college": request.form.get("college", "").strip(),
            "course": request.form.get("course", "").strip(),
            "reference_by": request.form.get("reference_by", "").strip(),
            "project": request.form.get("project", "").strip()
        }

        old_data = {
            "intern_name": intern["intern_name"],
            "age": str(intern["age"]),
            "contact": intern["contact"],
            "college": intern["college"],
            "course": intern["course"],
            "reference_by": intern["reference_by"],
            "project": intern["project"]
        }
        changes = {
            key: value
            for key, value in new_data.items()
            if value != old_data.get(key, "")
        }

        if not changes:
            flash("No changes detected. Your profile is already up to date.", "info")
        else:
            update_query = """
                UPDATE interns
                SET intern_name = %s,
                    age = %s,
                    contact = %s,
                    college = %s,
                    course = %s,
                    reference_by = %s,
                    project = %s
                WHERE email = %s
            """

            cur.execute(
                update_query,
                (
                    new_data["intern_name"],
                    new_data["age"],
                    new_data["contact"],
                    new_data["college"],
                    new_data["course"],
                    new_data["reference_by"],
                    new_data["project"],
                    email
                )
            )

            conn.commit()
            flash("Profile updated successfully.", "success")
            cur.execute(
                "SELECT * FROM interns WHERE email = %s",
                (email,)
            )
            intern = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "intern_dashboard.html",
        intern=intern,
        attendance=attendance
    )


@app.route('/api/attendance/latest')
def api_attendance_latest():
    if not session.get('intern_logged_in'):
        return jsonify({'error': 'not_logged_in'}), 401

    email = session.get('intern_email')
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT date, status FROM attendance WHERE intern_email = %s ORDER BY date DESC LIMIT 1",
        (email,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return jsonify({'date': None, 'status': None})
    latest_date = row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0])
    return jsonify({'date': latest_date, 'status': row[1]})


@app.route('/export_attendance/<reg_no>', methods=['POST'])
def export_attendance(reg_no):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute(
        "SELECT reg_no, intern_name, email FROM interns WHERE CAST(reg_no AS TEXT) = %s",
        (reg_no,)
    )
    intern = cur.fetchone()
    
    if not intern:
        cur.close()
        conn.close()
        flash('Intern not found for export', 'error')
        return redirect(url_for('admin_attendance', reg_no=reg_no))
    
    cur.execute(
        "SELECT date, status, marked_at FROM attendance WHERE intern_email = %s ORDER BY date ASC",
        (intern['email'],)
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    # Define exports directory
    exports_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(exports_dir, exist_ok=True)
    master_path = os.path.join(exports_dir, "attendance_master.xlsx")
    
    # Load or create workbook
    if os.path.exists(master_path):
        try:
            wb = load_workbook(master_path)
            ws = wb.active
        except Exception as e:
            print(f"Error loading workbook: {e}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            # Create headers with proper formatting
            headers = ["Reg No", "Intern Name", "Email", "Date", "Status", "Marked At"]
            ws.append(headers)
            # Make headers bold
            for cell in ws[1]:
                cell.font = Font(bold=True)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        headers = ["Reg No", "Intern Name", "Email", "Date", "Status", "Marked At"]
        ws.append(headers)
        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    # Collect all existing data from Excel (excluding header)
    all_records = {}  # key: (reg_no, date) -> [reg_no, name, email, date, status, marked_at]
    
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if row_cells[0]:  # If reg_no exists
            reg_val = str(row_cells[0]).strip()
            date_val = row_cells[3]
            
            # Handle date conversion
            if isinstance(date_val, datetime):
                date_key = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, str):
                date_key = date_val.strip()
            else:
                date_key = str(date_val).strip() if date_val else ""
            
            key = (reg_val, date_key)
            all_records[key] = list(row_cells)
    
    # Add/Update current intern's attendance
    for r in rows:
        date_obj = r['date']
        status = r['status']
        marked_at = r['marked_at']
        
        # Format date and time
        date_str = date_obj.strftime('%Y-%m-%d') if date_obj else ''
        marked_str = marked_at.strftime('%Y-%m-%d %H:%M:%S') if marked_at else ''
        
        key = (str(intern['reg_no']).strip(), date_str)
        
        # Update or add the record
        all_records[key] = [
            str(intern['reg_no']),
            intern['intern_name'],
            intern['email'],
            date_str,
            status,
            marked_str
        ]
    
    # Clear all rows except header
    ws.delete_rows(2, ws.max_row)
    
    # Sort records by reg_no first, then by date
    sorted_records = sorted(
        all_records.values(),
        key=lambda x: (
            str(x[0]),  # reg_no
            datetime.strptime(x[3], '%Y-%m-%d') if x[3] else datetime.min  # date
        )
    )
    
    # Write sorted data back to Excel
    for record in sorted_records:
        ws.append(record)
    
    # Auto-adjust column widths for better readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    try:
        wb.save(master_path)
        wb.close()
        print(f"Saved to: {master_path}")
        print(f"Total records in file: {len(sorted_records)}")
    except Exception as e:
        flash(f'Error saving attendance file: {str(e)}', 'error')
        return redirect(url_for('admin_attendance', reg_no=reg_no))
    
    # Check if download is requested
    download_value = request.form.get('download', '0')
    
    if download_value == '1':
        # Download the file
        try:
            return send_file(
                master_path,
                as_attachment=True,
                download_name="attendance_master.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            flash(f'Error downloading file: {str(e)}', 'error')
            return redirect(url_for('admin_attendance', reg_no=reg_no))
    else:
        # Just save without download
        flash(f'Attendance saved successfully! Total records: {len(sorted_records)}', 'success')
        return redirect(url_for('admin_attendance', reg_no=reg_no))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route('/_db_check')
def _db_check():
    """Temporary route: verifies DB connection and returns a small sample from interns."""
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT COUNT(*) AS cnt FROM interns")
        cnt = cur.fetchone()["cnt"]
        cur.execute("SELECT reg_no, intern_name, email FROM interns ORDER BY reg_no LIMIT 5")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"ok": True, "count": cnt, "sample": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


if __name__ == "__main__":
    try:
        ensure_admin_requests_table()
    except Exception:
        pass

    # Note: In Docker, Gunicorn runs this. 
    # This block only runs locally for development
    app.run(debug=False, host="0.0.0.0", port=5000)
